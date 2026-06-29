#!/usr/bin/env python3
"""
MedSync SKU Backfill
--------------------
Fetches MedSync products that are missing a SKU (NDC field), then looks up
the correct SKU from a Vetspire Products Report export.

Only updates products that already exist in MedSync — never adds new products.

Supported source workbooks (auto-detected):
  • Vetspire Products Report   — single sheet, columns vary (see VETSPIRE_COLS)
  • Q2 WH Inventory Count      — multiple pharmacy/vaccine sheets
  • MedSync De Novo Loading Order — single sheet, col B=name, col C=SKU

Usage (dry run):
  python3 sku_backfill.py --xlsx "Vetspire Products Report.xlsx"

Apply changes:
  SUPA_SERVICE_KEY="sb_secret_..." python3 sku_backfill.py --xlsx "Vetspire Products Report.xlsx" --apply
"""

import argparse
import json
import os
import re
import sys
import urllib.request
import urllib.error

try:
    import openpyxl
except ImportError:
    sys.exit("pip install openpyxl  then re-run")

# ── Config ─────────────────────────────────────────────────────────────────
SUPA_URL = "https://aemkdummdrmxtwrkggjw.supabase.co"
ANON_KEY = (
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9"
    ".eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImFlbWtkdW1tZHJteHR3cmtnZ2p3Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODAwOTQwNjEsImV4cCI6MjA5NTY3MDA2MX0"
    ".JzUojqfs9K6wOtrhjDnQ_knVU1wDvqR0MFH9z_r4G4s"
)

# Q2 inventory sheets: (name_col, sku_col), 0-indexed
SHEET_COLS = {
    "Pharmacy - General ":       (0, 2),
    "Pharmacy - Ocular ":        (0, 1),
    "Pharmacy - Preventatives ": (0, 1),
    "Pharmacy - Skin":           (0, 1),
    "Vaccines ":                 (0, 1),
    "IDEXX Labs ":               (0, 1),
    "IDEXX Consumables ":        (0, 1),
    "Diets":                     (0, 1),
    "Hospital Inventory":        (0, 4),
    "Stable Hospital Inventory": (0, 3),
}

DENOVO_SHEET   = "De Novo Loading Order"
VETSPIRE_SHEET = "Products"   # typical sheet name in Vetspire export

# ── Helpers ─────────────────────────────────────────────────────────────────

def _norm(s):
    s = str(s or "").lower().strip()
    s = re.sub(r"[^\w\s]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _sku_str(val):
    if val is None:
        return None
    if isinstance(val, float):
        return str(int(val)) if val == int(val) else str(val)
    return str(val).strip()


def _is_real_sku(s):
    """Reject empty, very short, or space-containing strings (those are names, not codes)."""
    if not s or len(s) < 4:
        return False
    if " " in s:
        return False
    return True


# ── Spreadsheet loaders ─────────────────────────────────────────────────────

def _find_header_row(ws, keywords):
    """Return (row_index, col_map) for the first row that contains all keywords.
    col_map: {keyword_lower: col_index}. Returns (None, {}) if not found.
    """
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True)):
        cells = [str(c or "").lower().strip() for c in row]
        found = {}
        for kw in keywords:
            for j, cell in enumerate(cells):
                if kw in cell:
                    found[kw] = j
                    break
        if len(found) == len(keywords):
            return i + 1, found  # 1-based row number
    return None, {}


def load_vetspire_report(ws):
    """Parse a Vetspire Products Report sheet.

    Vetspire exports vary slightly — we auto-detect column positions by header.
    Expected headers (case-insensitive, partial match OK): 'name', 'sku'
    """
    header_row, col_map = _find_header_row(ws, ["name", "sku"])
    if header_row is None:
        header_row, col_map = _find_header_row(ws, ["product", "sku"])
    if header_row is None:
        # Fallback: assume col 0 = name, col 1 = SKU
        print("  [warn] Vetspire sheet: could not find header row, assuming col 0=name col 1=SKU")
        name_col, sku_col = 0, 1
        data_start = 2
    else:
        name_col  = col_map.get("name", col_map.get("product"))
        sku_col   = col_map["sku"]
        data_start = header_row + 1

    mapping = {}
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        name    = row[name_col]  if len(row) > name_col  else None
        sku_raw = row[sku_col]   if len(row) > sku_col   else None
        if not name:
            continue
        sku_s = _sku_str(sku_raw)
        if not sku_s or not _is_real_sku(sku_s):
            continue
        key = _norm(name)
        if key and key not in ("medication", "item", "vaccine", "supplies"):
            mapping[key] = (sku_s, str(name).strip(), VETSPIRE_SHEET)
    return mapping


def load_spreadsheet(path):
    """Return dict: normalised_name → (sku, original_name, sheet).

    Priority order for auto-detection:
      1. Vetspire Products Report (sheet named 'Products' or single sheet with name+sku headers)
      2. De Novo Loading Order
      3. Q2 WH Inventory Count (multiple named sheets)
    """
    wb = openpyxl.load_workbook(path)
    mapping = {}

    # 1. Vetspire Products Report
    vetspire_ws = None
    if VETSPIRE_SHEET in wb.sheetnames:
        vetspire_ws = wb[VETSPIRE_SHEET]
    else:
        # Single-sheet workbook with 'name' and 'sku' columns → treat as Vetspire report
        if len(wb.sheetnames) == 1:
            ws = wb[wb.sheetnames[0]]
            hr, _ = _find_header_row(ws, ["name", "sku"])
            if hr is not None:
                vetspire_ws = ws

    if vetspire_ws is not None:
        print(f"  Detected: Vetspire Products Report (sheet: {vetspire_ws.title!r})")
        return load_vetspire_report(vetspire_ws)

    # 1b. Any single sheet with name+sku headers (catches 'Location XXXXX' exports)
    if len(wb.sheetnames) == 1:
        ws = wb[wb.sheetnames[0]]
        hr, _ = _find_header_row(ws, ["name", "sku"])
        if hr is None:
            # also try 'product' as the name column
            hr, _ = _find_header_row(ws, ["product", "sku"])
        if hr is not None:
            print(f"  Detected: Vetspire Products Report (sheet: {ws.title!r})")
            return load_vetspire_report(ws)

    # 2. De Novo Loading Order
    if DENOVO_SHEET in wb.sheetnames:
        print(f"  Detected: De Novo Loading Order")
        ws = wb[DENOVO_SHEET]
        for row in ws.iter_rows(min_row=5, values_only=True):
            row_num = row[0]
            name    = row[1]
            sku_raw = row[2]
            if not name or not isinstance(row_num, int):
                continue
            sku_s = _sku_str(sku_raw)
            if not sku_s or not _is_real_sku(sku_s):
                continue
            key = _norm(name)
            if key and key not in ("medication", "item", "vaccine", "supplies"):
                mapping[key] = (sku_s, str(name).strip(), DENOVO_SHEET)
        return mapping

    # 3. Q2 WH Inventory Count
    matched_sheets = [s for s in SHEET_COLS if s in wb.sheetnames]
    if matched_sheets:
        print(f"  Detected: Q2 WH Inventory Count ({len(matched_sheets)} sheets found)")
        for sheet, (icol, scol) in SHEET_COLS.items():
            if sheet not in wb.sheetnames:
                continue
            ws = wb[sheet]
            for row in ws.iter_rows(min_row=2, values_only=True):
                name = row[icol]
                sku  = row[scol]
                if not name or not sku:
                    continue
                sku_s = _sku_str(sku)
                if not sku_s:
                    continue
                key = _norm(name)
                if key and key not in ("medication", "item", "vaccine", "supplies"):
                    mapping[key] = (sku_s, str(name).strip(), sheet.strip())
        return mapping

    sys.exit(
        "ERROR: Could not identify workbook format.\n"
        "Expected one of:\n"
        "  • Vetspire Products Report (sheet named 'Products', or single sheet with Name/SKU columns)\n"
        "  • De Novo Loading Order\n"
        "  • Q2 WH Inventory Count\n"
        f"Sheets found: {wb.sheetnames}"
    )


# ── Supabase helpers ─────────────────────────────────────────────────────────

def supa_get(key, path):
    req = urllib.request.Request(
        SUPA_URL + path,
        headers={"apikey": key, "Authorization": f"Bearer {key}"}
    )
    with urllib.request.urlopen(req, timeout=30) as r:
        return json.loads(r.read())


def supa_patch(key, path, data):
    body = json.dumps(data).encode()
    req = urllib.request.Request(
        SUPA_URL + path,
        data=body,
        headers={
            "apikey":        key,
            "Authorization": f"Bearer {key}",
            "Content-Type":  "application/json",
            "Prefer":        "return=minimal",
        }
    )
    req.get_method = lambda: "PATCH"
    try:
        with urllib.request.urlopen(req, timeout=30) as r:
            return r.status
    except urllib.error.HTTPError as e:
        return e.code


# ── Manual overrides ─────────────────────────────────────────────────────────
# Keyed by normalised MedSync product name → correct SKU.
# Takes priority over any spreadsheet match.

MANUAL_OVERRIDES = {
    "vanguard bordetella vaccine oral sf":    "10014057",
    "vanguard rabies vaccine - 1 year":       "10016542",
    "purevax feline rabies vaccine - 1 year": "159730",
    "purevax feline rabies vaccine - 3 year": "159732",
}


# ── Matching ─────────────────────────────────────────────────────────────────

def find_sku(prod_name, sheet_map):
    """Return (sku, matched_name, source) or None."""
    key = _norm(prod_name)

    # 1. Manual override
    if key in MANUAL_OVERRIDES:
        return (MANUAL_OVERRIDES[key], prod_name + " [manual override]", "overrides")

    # 2. Exact normalised-key match
    match = sheet_map.get(key)
    if match:
        return match

    # 3. First-four-words fuzzy match (3-of-4 overlap)
    words = key.split()
    for skey, val in sheet_map.items():
        swords = skey.split()
        common = sum(1 for w in words[:4] if w in swords)
        if common >= 3:
            return val

    return None


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description="Backfill missing SKUs in MedSync products table")
    ap.add_argument("--xlsx", required=True, help="Path to source workbook (Vetspire Products Report recommended)")
    ap.add_argument("--apply", action="store_true", help="Write to Supabase (default: dry run)")
    args = ap.parse_args()

    service_key = os.environ.get("SUPA_SERVICE_KEY", "").strip()
    if args.apply and not service_key:
        sys.exit("ERROR: Set SUPA_SERVICE_KEY env var to apply changes.")

    read_key = service_key or ANON_KEY

    print("Loading spreadsheet …")
    sheet_map = load_spreadsheet(args.xlsx)
    print(f"  {len(sheet_map)} products with SKUs in workbook\n")

    print("Fetching MedSync products missing SKU …")
    all_products = []
    page_size = 1000
    offset = 0
    while True:
        batch = supa_get(
            read_key,
            f"/rest/v1/products?select=id,name,ndc&limit={page_size}&offset={offset}"
        )
        if isinstance(batch, dict) and "message" in batch:
            sys.exit(f"Supabase error: {batch}")
        all_products.extend(batch)
        if len(batch) < page_size:
            break
        offset += page_size

    total      = len(all_products)
    missing    = [p for p in all_products if not (p.get("ndc") or "").strip()]
    print(f"  {total} total products, {len(missing)} missing SKU\n")

    patches  = {}  # product_id → {ndc, product_name, matched_name, sheet}
    no_match = []

    for prod in missing:
        match = find_sku(prod["name"] or "", sheet_map)
        if match:
            sku, matched_name, sheet = match
            patches[prod["id"]] = {
                "ndc":          sku,
                "product_name": prod["name"],
                "matched_name": matched_name,
                "sheet":        sheet,
            }
        else:
            no_match.append(prod["name"] or "(unnamed)")

    # ── Report ───────────────────────────────────────────────────────────────
    print(f"{'='*70}")
    print(f"Products to update (SKU found):  {len(patches)}")
    print(f"No match (manual review needed): {len(no_match)}")
    print(f"{'='*70}\n")

    if patches:
        print("WILL ADD SKU:")
        for pid, info in patches.items():
            print(f"  {info['product_name']!r:60s} → {info['ndc']:15s}  [{info['sheet']}]")
        print()

    if no_match:
        print("NO MATCH FOUND:")
        for name in no_match:
            print(f"  {name!r}")
        print()

    if not args.apply:
        print("DRY RUN — pass --apply with SUPA_SERVICE_KEY set to write changes.")
        return

    # ── Apply ─────────────────────────────────────────────────────────────────
    print("Applying …")
    ok = err = 0
    for pid, info in patches.items():
        status = supa_patch(service_key, f"/rest/v1/products?id=eq.{pid}", {"ndc": info["ndc"]})
        if status in (200, 204):
            print(f"  ✓ {info['product_name']} → {info['ndc']}")
            ok += 1
        else:
            print(f"  ✗ {info['product_name']} → HTTP {status}")
            err += 1

    print(f"\nDone: {ok} updated, {err} errors.")


if __name__ == "__main__":
    main()
