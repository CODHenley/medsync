#!/usr/bin/env python3
"""
generate_sku_request.py
Generates sku_request.xlsx with all inventory items needing SKUs.
Run: python3 generate_sku_request.py
"""
import subprocess, sys

# Install openpyxl if needed
try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PRODUCTS = [
    # (Product Name, Manufacturer, Notes)
    ("Acepromazine Inj 10 mg/ml",                   "",        ""),
    ("Azithromycin 200 mg/5 ml Sus (30 ml bottle)", "",        ""),
    ("Azithromycin 250 mg tablets",                 "",        ""),
    ("Clevor Opth Solu",                            "Dechra",  "Ropinirole"),
    ("Elura (capromorelin)",                        "Elanco",  "Appetite stimulant"),
    ("Enrofloxacin Inj 22.7 mg/ml",                 "",        ""),
    ("Famotidine 20 mg tablets",                    "",        ""),
    ("Gabapentin Solu 50 mg/ml",                    "",        ""),
    ("Hill's Onc Feline 2.9 oz",                    "Hill's",  "UPC barcode, not NDC"),
    ("Mannitol 20% Inj (200 mg/ml)",                "",        ""),
    ("Marbofloxacin 25 mg tablets",                 "Dechra",  ""),
    ("Marbofloxacin 100 mg tablets",                "Dechra",  ""),
    ("Maropitant 160 mg tablets",                   "Zoetis",  ""),
    ("Mirtazapine 7.5 mg tablets",                  "",        ""),
    ("Nitro-Bid Ointment 2%",                       "",        ""),
    ("Ofloxacin 0.3% Otic Solu (5 ml)",             "",        ""),
    ("Oxytocin Inj 20 USP units/ml",                "",        ""),
    ("Pimobendan 2.5 mg tablets",                   "",        ""),
    ("Royal Canin Feline SO (3 oz)",                "Royal Canin", "UPC barcode, not NDC"),
    ("Zenalpha 0.5 mg/ml",                          "Dechra",  ""),
]

wb = Workbook()
ws = wb.active
ws.title = "SKU Request"

# ── Styles ────────────────────────────────────────────────────────────────────
navy    = "1C2B4A"
purple  = "F5F3FA"
border_color = "E2DFF0"

hdr_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
hdr_fill  = PatternFill("solid", fgColor=navy)
row_fill  = PatternFill("solid", fgColor="FFFFFF")
alt_fill  = PatternFill("solid", fgColor=purple)
thin      = Side(style="thin", color=border_color)
cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── Header row ────────────────────────────────────────────────────────────────
headers = ["Product Name", "Manufacturer", "NDC / SKU (please fill in)", "Notes"]
col_widths = [52, 16, 28, 22]

for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font      = hdr_font
    cell.fill      = hdr_fill
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    cell.border    = cell_border
    ws.column_dimensions[get_column_letter(col)].width = w

ws.row_dimensions[1].height = 22

# ── Data rows ─────────────────────────────────────────────────────────────────
for i, (name, mfr, notes) in enumerate(PRODUCTS, 2):
    fill = alt_fill if i % 2 == 0 else row_fill
    values = [name, mfr, "", notes]
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=i, column=col, value=val)
        cell.font      = Font(name="Calibri", size=10)
        cell.fill      = fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border    = cell_border
    ws.row_dimensions[i].height = 16

# ── Freeze header, auto-filter ────────────────────────────────────────────────
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:D{len(PRODUCTS)+1}"

out = "sku_request.xlsx"
wb.save(out)
print(f"Saved → {out}  ({len(PRODUCTS)} products)")
